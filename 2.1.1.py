import csv
from openpyxl import Workbook
from openpyxl.styles import Border, Side, Alignment, Font
from openpyxl.utils import get_column_letter
import openpyxl

designation = 0
salaryfrom = 1
salaryto = 2
salary = 3
areaname = 4
pubtime = 5

currency_to_rub = {
    "AZN": 35.68,
    "BYR": 23.91,
    "EUR": 59.90,
    "GEL": 21.74,
    "KGS": 0.76,
    "KZT": 0.13,
    "RUR": 1,
    "UAH": 1.64,
    "USD": 60.66,
    "UZS": 0.0055,
}

class Report:
    def __init__(self, fileN, naming):
        self.filename = fileN
        self.name = naming
        self.years = list(range(2007, 2023))
        self.sums = {}
        self.length = {}
        self.sumsCur = {}
        self.lengthCur = {}
        self.cities = []
        self.citySum = {}
        self.cityLength = {}
        self.vacLength = 0
        self.ansSums = {}
        self.part = {}
        self.csv_reader()
        self.calculate_file()
        self.workBook = Workbook()

    def csv_reader(self):
        point = False
        with open(self.filename, encoding="utf-8") as file:
            reader = csv.reader(file)
            for str in reader:
                if not point:
                    point = True
                    designation = str.index("name")
                    salaryfrom = str.index("salary_from")
                    salaryto = str.index("salary_to")
                    salary = str.index("salary_currency")
                    areaname = str.index("area_name")
                    pubtime = str.index("published_at")
                else:
                    str = str.copy()
                    if all(str):
                        curYear = int(str[pubtime].split("-")[0])
                        curSalary = (int(float(str[salaryto])) + int(float(str[salaryfrom]))) * currency_to_rub[
                            str[salary]] // 2
                        curTitle = str[designation]
                        curCity = str[areaname]
                        self.sums[curYear] = self.sums.get(curYear, 0) + curSalary
                        self.length[curYear] = self.length.get(curYear, 0) + 1
                        if profession in curTitle:
                            self.sumsCur[curYear] = self.sumsCur.get(curYear, 0) + curSalary
                            self.lengthCur[curYear] = self.lengthCur.get(curYear, 0) + 1
                        if curCity not in self.cities:
                            self.cities.append(curCity)
                        self.citySum[curCity] = self.citySum.get(curCity, 0) + curSalary
                        self.cityLength[curCity] = self.cityLength.get(curCity, 0) + 1
                        self.vacLength += 1

    def calculate_file(self):
        for i in self.years:
            if self.sums.get(i, None):
                self.sums[i] = int(self.sums[i] // self.length[i])
            if self.sumsCur.get(i, None):
                self.sumsCur[i] = int(self.sumsCur[i] // self.lengthCur[i])

        for i in self.cities:
            self.citySum[i] = int(self.citySum[i] // self.cityLength[i])
        required_cities = [city for city in self.cities if self.cityLength[city] >= self.vacLength // 100]
        self.ansSums = {key: self.citySum[key] for key in
                              sorted(required_cities, key=lambda x: self.citySum[x], reverse=True)[:10]}
        self.part = {key: float("{:.4f}".format(self.cityLength[key] / self.vacLength)) for key in sorted(required_cities, key=lambda x: self.cityLength[x] / self.vacLength, reverse=True)[:10]}

    def print_file(self):
        print("Динамика уровня зарплат по годам:", self.sums)
        print("Динамика количества вакансий по годам:", self.length)

        if not len(self.sumsCur):
            self.sumsCur[2022] = 0

        print("Динамика уровня зарплат по годам для выбранной профессии:", self.sumsCur)

        if not len(self.lengthCur):
            self.lengthCur[2022] = 0

        print("Динамика количества вакансий по годам для выбранной профессии:", self.lengthCur)
        print("Уровень зарплат по городам (в порядке убывания):", self.ansSums)
        print("Доля вакансий по городам (в порядке убывания):", self.part)

    def generate_xls(self):
        self.yearsStatSheet = self.workBook.create_sheet(title="Статистика по годам")
        self.citiesStatSheet = self.workBook.create_sheet(title="Статистика по городам")
        self.workBook.remove(self.workBook["Sheet"])
        side = Side(border_style='thin', color="000000")
        self.border = Border(right=side, top=side, bottom=side, left=side)
        self.alignment = Alignment(horizontal='left')
        self.dataAlignment = Alignment(horizontal='right')
        self.citiesStatSheet["a1"] = 12
        self.report_years()
        self.report_cities()
        self.suitable_cells()
        self.workBook.save('report.xlsx')

    def report_years(self):
        titles = ["Год", "Средняя зарплата", "Средняя зарплата - " + self.name,
                   "Количество вакансий", "Количество вакансий - " + self.name]
        self.set_headers(self.yearsStatSheet, titles)

        matrix = []
        for value in range(len(self.sums)):
            data = list(self.sums.keys())[value]
            append = [data, self.sums[data], self.sumsCur[data], self.length[data],
                          self.lengthCur[data]]
            matrix.append(append)
        self.complete_matrix(self.yearsStatSheet, matrix, offset=(0, 1))

    def complete_matrix(self, sheet, matrix, offset=(0, 0)):
        for i in range(len(matrix)):
            for j in range(len(matrix[0])):
                address = f"{get_column_letter(j + 1 + offset[0])}{i + 1 + offset[1]}"
                sheet[address] = matrix[i][j]
                sheet[address].border = self.border
                sheet[address].alignment = self.dataAlignment
                sheet.column_dimensions[get_column_letter(j + 1)].auto_size = 1

    def set_headers(self, sheet, headers, offset=(0, 0)):
        for i in range(0, len(headers)):
            address = f"{get_column_letter(i + 1 + offset[0])}{1 + offset[1]}"
            sheet[address] = headers[i]
            sheet[address].border = self.border
            sheet[address].alignment = self.alignment
            sheet[address].font = Font(bold=True)
            sheet.column_dimensions[get_column_letter(i + 1)].auto_size = 1

    def suitable_cells(self):
        for sheetName in self.workBook.sheetnames:
            sheetWB = self.workBook[sheetName]
            for i in range(1, sheetWB.max_column + 1):
                width = None
                for j in range(1, sheetWB.max_row + 1):
                    value = sheetWB[f"{get_column_letter(i)}{j}"].value
                    if value is not None and (width is None or len(str(value)) > width):
                        width = len(str(value))
                if width is not None:
                    sheetWB.column_dimensions[f"{get_column_letter(i)}"].width = width + 2
                else:
                    sheetWB.column_dimensions[f"{get_column_letter(i)}"].width = + 2

    def report_cities(self):
        salaryLevel = ["Город", "Уровень зарплат"]
        shareOfVacancies = ["Город", "Доля вакансий"]
        self.set_headers(self.citiesStatSheet, salaryLevel)
        self.set_headers(self.citiesStatSheet, shareOfVacancies, (3, 0))
        self.dataAlignment = Alignment(horizontal='left')
        self.complete_matrix(self.citiesStatSheet, [[i] for i in self.ansSums.keys()], offset=(0, 1))
        max = {key: f"{(val * 10000) // 1 / 100}%" for key, val in self.part.items()}
        self.complete_matrix(self.citiesStatSheet, [[i] for i in list(max.keys())], offset=(3, 1))
        self.dataAlignment = Alignment(horizontal='right')
        self.complete_matrix(self.citiesStatSheet, [[i] for i in list(self.ansSums.values())], offset=(1, 1))
        self.complete_matrix(self.citiesStatSheet, [[i] for i in list(max.values())], offset=(4, 1))


file_name = input("Введите название файла: ")
profession = input("Введите название профессии: ")
report = Report(file_name, profession)
report.print_file()
report.generate_xls()   
