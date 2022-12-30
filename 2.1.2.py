import csv
from openpyxl import Workbook
from openpyxl.styles import Border, Side, Alignment, Font
from openpyxl.utils import get_column_letter
import matplotlib.pyplot as plt
import matplotlib
import numpy as np
from jinja2 import Environment, FileSystemLoader
import pdfkit
import os

title = 0
salary_minimum = 1
salary_maximum = 2
salary = 3
area = 4
published = 5

currency_to_rub = {
    "AZN": 35.68,
    "BYR": 23.91,
    "EUR": 59.90,
    "GEL": 21.74,
    "KGS": 0.76,
    "KZT": 0.13,
    "RUR": 1,
    "UAH": 1.64,
    "USD": 60.66,
    "UZS": 0.0055,
}

class Report:
    def __init__(self, fileName, name):
        self.filename = fileName
        self.name = name
        self.years = list(range(2007, 2023))
        self.years_sums = {}
        self.years_length = {}
        self.years_sums_cur = {}
        self.years_length_cur = {}
        self.cities = []
        self.cities_sums = {}
        self.cities_length = {}
        self.vacancies_length = 0
        self.ansCitiesSums = {}
        self.citiesPartitions = {}
        self.csv_reader()
        self.calculate_file()
        self.Wb = Workbook()

    def csv_reader(self):
        flag = False
        with open(self.filename, encoding="utf-8") as file:
            reader = csv.reader(file)
            for data in reader:
                if not flag:
                    flag = True
                    title = data.index("name")
                    salary_minimum = data.index("salary_from")
                    salary_maximum = data.index("salary_to")
                    salary = data.index("salary_currency")
                    area = data.index("area_name")
                    published = data.index("published_at")
                else:
                    data = data.copy()
                    if all(data):
                        curr_year = int(data[published].split("-")[0])
                        curr_salary = (int(float(data[salary_maximum])) + int(float(data[salary_minimum]))) * currency_to_rub[data[salary]] // 2
                        curr_name = data[title]
                        curr_city = data[area]
                        self.years_sums[curr_year] = self.years_sums.get(curr_year, 0) + curr_salary
                        self.years_length[curr_year] = self.years_length.get(curr_year, 0) + 1
                        if profession in curr_name:
                            self.years_sums_cur[curr_year] = self.years_sums_cur.get(curr_year, 0) + curr_salary
                            self.years_length_cur[curr_year] = self.years_length_cur.get(curr_year, 0) + 1
                        if curr_city not in self.cities:
                            self.cities.append(curr_city)
                        self.cities_sums[curr_city] = self.cities_sums.get(curr_city, 0) + curr_salary
                        self.cities_length[curr_city] = self.cities_length.get(curr_city, 0) + 1
                        self.vacancies_length += 1

    def calculate_file(self):
        for i in self.years:
            if self.years_sums.get(i, None):
                self.years_sums[i] = int(self.years_sums[i] // self.years_length[i])
            if self.years_sums_cur.get(i, None):
                self.years_sums_cur[i] = int(self.years_sums_cur[i] // self.years_length_cur[i])

        for j in self.cities:
            self.cities_sums[j] = int(self.cities_sums[j] // self.cities_length[j])
        interesting_cities = [city for city in self.cities if self.cities_length[city] >= self.vacancies_length // 100]
        self.ansCitiesSums = {key: self.cities_sums[key] for key in sorted(interesting_cities, key=lambda x: self.cities_sums[x], reverse=True)[:10]}
        self.citiesPartitions = {key: float("{:.4f}".format(self.cities_length[key] / self.vacancies_length)) for key in sorted(interesting_cities, key=lambda x: self.cities_length[x] / self.vacancies_length,
                                         reverse=True)[:10]}

    def print_file(self):
        print("Динамика уровня зарплат по годам:", self.years_sums)
        print("Динамика количества вакансий по годам:", self.years_length)
        if not len(self.years_sums_cur):
            self.years_sums_cur[2022] = 0
        print("Динамика уровня зарплат по годам для выбранной профессии:", self.years_sums_cur)
        if not len(self.years_length_cur):
            self.years_length_cur[2022] = 0
        print("Динамика количества вакансий по годам для выбранной профессии:", self.years_length_cur)
        print("Уровень зарплат по городам (в порядке убывания):", self.ansCitiesSums)
        print("Доля вакансий по городам (в порядке убывания):", self.citiesPartitions)

    def generate_xls(self):
        self.years_stat_sheet = self.Wb.create_sheet(title="Статистика по годам")
        self.cities_stat_sheet = self.Wb.create_sheet(title="Статистика по городам")
        self.Wb.remove(self.Wb["Sheet"])
        side = Side(border_style='thin', color="000000")
        self.border = Border(right=side, top=side, bottom=side, left=side)
        self.header_alignment = Alignment(horizontal='left')
        self.data_alignment = Alignment(horizontal='right')
        self.cities_stat_sheet["a1"] = 12
        self.report_years()
        self.report_cities()
        self.suitable_cells()
        self.Wb.save('report.xlsx')

    def report_years(self):
        headers = ["Год", "Средняя зарплата", "Средняя зарплата - " + self.name,
                   "Количество вакансий", "Количество вакансий - " + self.name]
        self.set_headers(self.years_stat_sheet, headers)

        matrix = []
        for row in range(len(self.years_sums)):
            key = list(self.years_sums.keys())[row]
            appendable = [key, self.years_sums[key], self.years_sums_cur[key], self.years_length[key],
                          self.years_length_cur[key]]
            matrix.append(appendable)

        self.complete_matrix(self.years_stat_sheet, matrix, offset=(0, 1))

    def complete_matrix(self, sheet, matrix, offset=(0, 0)):
        for row in range(len(matrix)):
            for col in range(len(matrix[0])):
                address = f"{get_column_letter(col + 1 + offset[0])}{row + 1 + offset[1]}"
                sheet[address] = matrix[row][col]
                sheet[address].border = self.border
                sheet[address].alignment = self.data_alignment
                sheet.column_dimensions[get_column_letter(col + 1)].auto_size = 1

    def set_headers(self, sheet, headers, offset=(0, 0)):
        for col in range(0, len(headers)):
            address = f"{get_column_letter(col + 1 + offset[0])}{1 + offset[1]}"
            sheet[address] = headers[col]
            sheet[address].border = self.border
            sheet[address].alignment = self.header_alignment
            sheet[address].font = Font(bold=True)
            sheet.column_dimensions[get_column_letter(col + 1)].auto_size = 1

    def suitable_cells(self):
        for sheet_name in self.Wb.sheetnames:
            sheet = self.Wb[sheet_name]
            for col in range(1, sheet.max_column + 1):
                width = None
                for row in range(1, sheet.max_row + 1):
                    value = sheet[f"{get_column_letter(col)}{row}"].value
                    if value is not None and (width is None or len(str(value)) > width):
                        width = len(str(value))
                if width is not None:
                    sheet.column_dimensions[f"{get_column_letter(col)}"].width = width + 2
                else:
                    sheet.column_dimensions[f"{get_column_letter(col)}"].width = + 2

    def report_cities(self):
        headers_salarylevel = ["Город", "Уровень зарплат"]
        headers_percent = ["Город", "Доля вакансий"]
        self.set_headers(self.cities_stat_sheet, headers_salarylevel)
        self.set_headers(self.cities_stat_sheet, headers_percent, (3, 0))

        self.data_alignment = Alignment(horizontal='left')
        self.complete_matrix(self.cities_stat_sheet, [[i] for i in self.ansCitiesSums.keys()], offset=(0, 1))
        matrix = {key: f"{(val * 10000) // 1 / 100}%" for key, val in self.citiesPartitions.items()}
        self.complete_matrix(self.cities_stat_sheet, [[i] for i in list(matrix.keys())], offset=(3, 1))
        self.data_alignment = Alignment(horizontal='right')
        self.complete_matrix(self.cities_stat_sheet, [[i] for i in list(self.ansCitiesSums.values())], offset=(1, 1))
        self.complete_matrix(self.cities_stat_sheet, [[i] for i in list(matrix.values())], offset=(4, 1))

    def generate_img(self):
        matplotlib.rc("font", size=8)
        fig, ((ax1, ax2), (ax3, ax4)) = plt.subplots(nrows=2, ncols=2)
        width = 0.3
        x = np.arange(len(self.years_sums.keys()))
        payment1 = ax1.bar(x - width / 2, self.years_sums.values(), width, label="средняя з/п")
        payment2 = ax1.bar(x + width / 2, self.years_sums_cur.values(), width, label=f"з/п {self.name}")

        ax1.grid(True, axis="y")
        ax1.set_title("Уровень зарплат по годам")
        ax1.set_xticks(np.arange(len(self.years_sums.keys())), self.years_sums.keys(), rotation=90)
        ax1.bar_label(payment1, fmt="")
        ax1.bar_label(payment2, fmt="")
        ax1.legend(prop={"size": 6})

        ax2.grid(True, axis="y")
        ax2.set_title("Количество вакансий по годам")
        x = np.arange(len(self.years_sums.keys()))
        ax2.set_xticks(x, self.years_sums.keys(), rotation=90)
        vac1 = ax2.bar(x - width / 2, self.years_sums.values(), width, label="Количество вакансий")
        vac2 = ax2.bar(x + width / 2, self.years_sums_cur.values(), width, label=f"Количество вакансий\n{self.name}")
        ax2.bar_label(vac1, fmt="")
        ax2.bar_label(vac2, fmt="")
        ax2.legend(prop={"size": 6})

        ax3.grid(True, axis="x")
        y = np.arange(len(list(self.ansCitiesSums.keys())))
        ax3.set_yticks(y, map(lambda s: s.replace(" ", "\n").replace("-", "\n"), self.ansCitiesSums.keys()))
        ax3.invert_yaxis()
        ax3.barh(y, self.ansCitiesSums.values())
        ax3.set_title("Уровень зарплат по городам")

        ax4.set_title("Доля вакансий по городам")
        other = 1 - sum(self.citiesPartitions.values())
        ax4.pie([other] + list(self.citiesPartitions.values()),
                labels=["Другие"] + list(self.citiesPartitions.keys()), startangle=0)

        fig.tight_layout(pad=0.4, w_pad=0.5, h_pad=1.0)
        plt.savefig("graph.png")

    def generate_pdf(self):
        self.generate_img()
        env = Environment(loader=FileSystemLoader('.'))
        template = env.get_template("pdf_template.html")
        pt = os.path.abspath("graph.png")

        statistics = {}
        for k in self.years:
            if self.years_sums.get(k, None) is not None:
                statistics[k] = [self.years_sums[k], self.years_sums_cur[k], self.years_length[k],
                                 self.years_length_cur[k]]

        file_template = template.render(
            {"plot": pt,
             "name": self.name,
             "years_stat": statistics,
             "cities_sum": self.ansCitiesSums,
             "cities_part": {key: ((val * 10000) // 1) / 100 for key, val in self.citiesPartitions.items()}
             })
        path = pdfkit.configuration(wkhtmltopdf=r'C:\Program Files\wkhtmltopdf\bin\wkhtmltopdf.exe')
        pdfkit.from_string(file_template, "report.pdf", configuration=path, options={"enable-local-file-access": ""})


fileName = input("Введите название файла: ")
profession = input("Введите название профессии: ")
report = Report(fileName, profession)
report.print_file()
report.generate_pdf()
